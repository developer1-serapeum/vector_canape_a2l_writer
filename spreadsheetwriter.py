import re
from openpyxl import Workbook
from openpyxl import load_workbook


"""
The goal of this script write to an excel sheet
with custom columns  
"""

class SpreadSheetWriter():

    def __init__(self):
        pass

    def writeToFile(self, file_out):
        """
        Write to the file where each line has exactly one warning
        """
        with open(file_out, 'w') as f:
            for line in new_lines:
                f.write("%s\n" % line)


    def writeToExcel(self, fileName, column1, column2, column3, column4):
        """
        Takes 4 input arrarys and writes them as columns in an excell sheet
        """
        workbook = Workbook()
        sheet = workbook.active
        #sheet["A1"] = "Hello World !"
        for row in range(1, len(column1)+1):
            sheet.cell(row=row, column=1).value = column1[row-1]
            sheet.cell(row=row, column=2).value = column2[row-1]
            sheet.cell(row=row, column=3).value = column3[row-1]
            sheet.cell(row=row, column=4).value = column4[row-1]
        
        workbook.save(fileName)


    def writeToExcelOneColumn(self, fileName, column1):
        """
        A concrete function to create a sheet with one column only
        """
        workbook = load_workbook(fileName)
        sheet = workbook["All Requirements"]
        selectedColumn = sheet.max_column + 1
        for row in range(1, len(column1)+1):
            sheet.cell(row=row, column=selectedColumn).value = column1[row-1]
        
        workbook.save(fileName)


    def findPattern(self):
        """
        [For debug only] Finds a desired pattern within a text file
        and writes it to an output file
        """
        file_in = 'compiler_in.txt'
        file_out = 'compiler_warning_ids.txt'
        pattern = re.compile(r"warning\s(#.+-.):")
        
        match_count = 0
        found_instances = []
        for i, line in enumerate(open(file_in)):
            for match in re.finditer(pattern, line):
                print ('Found on line %s: %s' % (i+1, match.group()) )
                found_instances.append(match.group(1))
                match_count += 1

        print("The pattern was found ({}) times".format(match_count))

        with open(file_out, 'w') as f:
            for line in found_instances:
                f.write("%s\n" % line)


    def print_rows(self):
        """
        [For debug only] It makes it easier to print all of your spreadsheet values by
        just calling print_rows().
        """
        workbook = Workbook()
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            print(row)




    def initializedSRTM(self, fileName):

        workbook = Workbook()
        workbook.create_sheet("Analysis Results")
        workbook.create_sheet("All Requirements")
        workbook.save(fileName)
