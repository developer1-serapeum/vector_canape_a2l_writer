import re
from openpyxl import Workbook
from openpyxl import load_workbook


"""
The goal of this script is to read parameters from excel sheet
"""

class SpreadSheetReader():

    def __init__(self):
        pass

    def readParameter(self, fileName, tabName, paremeterName):
        """
        Read a single parameter
        """
        wb = load_workbook(filename = fileName)
        sheet_ranges = wb[tabName]
        print(sheet_ranges['A'+paremeterName].value)
        
    def readAllParameters(self, fileName, tabName, column):
        """
        Read all parameters in the specified colum
        """
        wb = load_workbook(filename = fileName)
        sheet_ranges = wb[tabName]
        parameters = []

        for row in range(2,sheet_ranges.max_row+1):  
            cell_name = "{}{}".format(column, row)
            cell_value = sheet_ranges[cell_name].value 
            if (cell_value != None):
                parameters.append(cell_value)

        return parameters

        